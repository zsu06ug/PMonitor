import re
import os
import json
import hashlib
import pandas as pd

from kivy.metrics import dp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner
from kivy.uix.filechooser import FileChooserListView
from kivy.clock import Clock
from kivy.graphics import Color, RoundedRectangle
from kivymd.app import MDApp
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.label import MDLabel
from kivymd.uix.card import MDCard, MDSeparator
from kivymd.uix.textfield import MDTextField
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.toast import toast
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.navigationdrawer import MDNavigationDrawer, MDNavigationLayout
from kivymd.uix.list import (
    MDList,
    OneLineListItem,
    OneLineIconListItem,
    IconLeftWidget,
)
from kivymd.uix.dialog import MDDialog
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.screen import MDScreen
from openpyxl import load_workbook
from kivy.uix.anchorlayout import AnchorLayout
from kivymd.uix.button import MDIconButton


def extract_number(value):
    if value is None or str(value).strip() == "":
        return None
    cleaned = str(value).replace(" ", "").replace(",", ".")
    cleaned = re.sub(r"[^\d.]", "", cleaned)
    try:
        return float(cleaned)
    except ValueError:
        return None


# ---------- ЕКРАН ВИБОРУ ФАЙЛУ ----------
class FileScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(
            orientation="vertical",
            padding=[dp(12), dp(20), dp(12), dp(20)],
            spacing=dp(20),
            **kwargs,
        )

        # Контейнер з темним фоном спеціально для файлового менеджера,
        # щоб білий текст стандартного FileChooser завжди було добре видно
        self.fc_wrapper = BoxLayout(orientation="vertical")
        with self.fc_wrapper.canvas.before:
            Color(0.12, 0.12, 0.12, 1)  # Темно-сірий колір фону
            self.fc_bg = RoundedRectangle(radius=[12])
        self.fc_wrapper.bind(pos=self._update_bg, size=self._update_bg)

        self.filechooser = FileChooserListView(
            path="/storage/emulated/0/", filters=["*.xlsx"]
        )
        self.fc_wrapper.add_widget(self.filechooser)
        self.add_widget(self.fc_wrapper)

        self.open_btn = MDRaisedButton(
            text="ВІДКРИТИ ОБРАНИЙ ФАЙЛ",
            size_hint=(1, None),
            height=dp(60),
            elevation=2,
            on_release=self.load_excel,
        )
        self.add_widget(self.open_btn)

    def _update_bg(self, instance, value):
        self.fc_bg.pos = instance.pos
        self.fc_bg.size = instance.size

    def load_excel(self, instance):
        if not self.filechooser.selection:
            return

        path = self.filechooser.selection[0]
        app = MDApp.get_running_app()
        app.source_file_path = path
        app.load_data_from_source()


# ---------- КАРТКА ТОВАРУ ----------
class ProductCard(MDCard):
    def __init__(
        self,
        index,
        tm,
        name,
        article,
        price,
        comp_price,
        competitor_name,
        promo_price,
        promo_checked,
        note_text,
        uom,
        segment,
        **kwargs,
    ):
        super().__init__(
            orientation="vertical",
            padding=[dp(12), dp(10), dp(12), dp(10)],
            spacing=dp(2),
            size_hint_y=None,
            height=dp(280),
            radius=[12],
            line_color=(0.5, 0.5, 0.5, 0.5),
            line_width=1,
            **kwargs,
        )

        self.index = index
        self.our_price_float = extract_number(price)

        display_price = (
            f"{self.our_price_float:.2f}".replace(".", ",")
            if self.our_price_float is not None
            else str(price).replace(".", ",")
        )
        clean_article = (
            str(article).split(".")[0] if "." in str(article) else str(article)
        )
        display_uom = (
            str(uom).strip()
            if str(uom).strip() not in ["", "nan", "None"]
            else ""
        )
        display_segment = (
            str(segment).strip()
            if str(segment).strip() not in ["", "nan", "None"]
            else "Не вказано"
        )
        display_tm = (
            str(tm).strip()
            if str(tm).strip() not in ["", "nan", "None"]
            else "Без ТМ"
        )

        # 1. ТМ
        self.add_widget(
            MDLabel(
                text=f"[color=#d4a000][b]{display_tm}[/b][/color]",
                markup=True,
                font_style="Caption",
                size_hint_y=None,
                height=dp(16),
                halign="left",
            )
        )

        # 2. Назва
        self.add_widget(
            MDLabel(
                text=f"[b]{name}[/b]",
                markup=True,
                font_style="Subtitle1",
                theme_text_color="Primary",
                size_hint_y=None,
                height=dp(44),
                halign="left",
            )
        )

        # 3. Інфо
        display_comp_name = (
            str(competitor_name).strip()
            if str(competitor_name).strip() not in ["", "nan", "None"]
            else "Не вказано"
        )
        uom_part = f" / {display_uom}" if display_uom else ""

        info_text = (
            f"Арт: [color=#777777]{clean_article}{uom_part}[/color]\n"
            f"Сегм: [color=#777777]{display_segment}[/color]\n"
            f"Конк: [color=#2288aa][b]{display_comp_name}[/b][/color]"
        )
        self.add_widget(
            MDLabel(
                text=info_text,
                markup=True,
                font_style="Caption",
                size_hint_y=None,
                height=dp(46),
                halign="left",
            )
        )

        # 4. Ціни
        promo_float = extract_number(promo_price)
        display_promo = (
            f"{promo_float:.2f}".replace(".", ",")
            if promo_float is not None
            else str(promo_price).replace(".", ",")
        )
        if display_promo.strip() == "" or display_promo == "nan":
            display_promo = "-"

        prices_text = f"Ціна: [b]{display_price}[/b]\nАкційна ціна: [color=#e53935]{display_promo}[/color]"
        self.add_widget(
            MDLabel(
                text=prices_text,
                markup=True,
                font_style="Body2",
                theme_text_color="Secondary",
                size_hint_y=None,
                height=dp(34),
                halign="left",
            )
        )

        # --- 5. ОБ'ЄДНАНИЙ РЯДОК: Ціна конкурента + Чекбокс Акції ---
        action_row = BoxLayout(
            orientation="horizontal",
            size_hint_y=None,
            height=dp(50),
            spacing=dp(12),
            padding=(0, 0, dp(6), 0),  # невеликий відступ справа
        )

        # Поле вводу ціни
        initial_comp_price = ""
        comp_price_float = extract_number(comp_price)

        if comp_price_float is not None:
            initial_comp_price = f"{comp_price_float:.2f}".replace(".", ",")

        self.input_price = MDTextField(
            hint_text="Ціна конкурента",
            text=initial_comp_price,
            mode="line",
            size_hint_x=1,  # поле займає весь вільний простір
        )

        self.input_price.bind(focus=self.on_price_focus)
        self.input_price.bind(text=self.on_price_text_change)

        self.update_color(self.input_price.text)

        action_row.add_widget(self.input_price)

        # Контейнер чекбокса
        raw_promo = str(promo_checked).strip().lower()
        is_promo_active = raw_promo in [
            "да",
            "да ",
            " да",
            "da",
            "yes",
            "true",
            "1",
        ]

        chk_wrap = BoxLayout(
            orientation="horizontal",
            size_hint=(None, 1),
            width=dp(110),  # фіксована ширина блоку справа
            spacing=dp(4),
            pos_hint={"center_y": 0.5},
        )

        self.chk = MDCheckbox(
            active=is_promo_active,
            size_hint=(None, None),
            size=(dp(36), dp(36)),
            pos_hint={"center_y": 0.5},
        )

        self.chk.bind(active=self.on_checkbox_active)

        chk_label = MDLabel(
            text="Акція", font_style="Body2", halign="left", valign="middle"
        )

        chk_wrap.add_widget(self.chk)
        chk_wrap.add_widget(chk_label)

        action_row.add_widget(chk_wrap)

        self.add_widget(action_row)

        # 6. Примітка
        note_wrap = BoxLayout(size_hint_y=None, height=dp(46))
        self.input_note = MDTextField(
            hint_text="Примітка",
            text=(
                str(note_text)
                if str(note_text).strip() not in ["", "nan", "None"]
                else ""
            ),
            mode="line",
        )
        self.input_note.bind(text=self.on_note_text_change)
        note_wrap.add_widget(self.input_note)
        self.add_widget(note_wrap)

    def on_checkbox_active(self, checkbox, value):
        app = MDApp.get_running_app()
        status = "Да" if value else ""
        if app.dataframe.at[self.index, "Акция"] != status:
            app.dataframe.at[self.index, "Акция"] = status
            app.save_change_to_service_file(
                self.index, "Incremented_Status_Dummy", "1"
            )
            app.save_change_to_service_file(self.index, "Акция", status)

    def on_price_text_change(self, instance, text):
        self.update_color(text)
        app = MDApp.get_running_app()
        val_float = extract_number(text)
        new_val = (
            "" if val_float is None else f"{val_float:.2f}".replace(".", ",")
        )

        if str(app.dataframe.at[self.index, "Цена конкурента"]) != new_val:
            app.dataframe.at[self.index, "Цена конкурента"] = new_val
            app.save_change_to_service_file(
                self.index, "Цена конкурента", new_val
            )

    def on_price_focus(self, instance, focus):
        if not focus:
            val_float = extract_number(instance.text)
            if val_float is not None:
                instance.text = f"{val_float:.2f}".replace(".", ",")

    def on_note_text_change(self, instance, text):
        app = MDApp.get_running_app()
        new_note = str(text).strip()
        if str(app.dataframe.at[self.index, "Примечание"]) != new_note:
            app.dataframe.at[self.index, "Примечание"] = new_note
            app.save_change_to_service_file(self.index, "Примечание", new_note)

    def update_color(self, text_val):
        app = MDApp.get_running_app()

        # Залежно від теми ставимо базовий колір (дефолтний)
        if app.theme_cls.theme_style == "Dark":
            base_color = [1, 1, 1, 1]  # Білий для темної теми
        else:
            base_color = [0, 0, 0, 0.6]  # Темно-сірий/чорний для світлої теми

        color = base_color

        c_price = extract_number(text_val)
        if (
            c_price is not None
            and self.our_price_float is not None
            and self.our_price_float > 0
        ):
            diff_percent = (
                abs(c_price - self.our_price_float) / self.our_price_float * 100
            )
            if diff_percent <= 5:
                color = [0.2, 0.6, 0.2, 1]
            elif diff_percent <= 10:
                color = [0.8, 0.6, 0.0, 1]
            else:
                color = [0.8, 0.2, 0.2, 1]

        self.input_price.text_color_normal = color
        self.input_price.text_color_focus = color
        self.input_price.line_color_normal = color
        self.input_price.line_color_focus = color


# ---------- ГОЛОВНИЙ ЕКРАН ----------
class MainScreen(BoxLayout):
    def __init__(self, rows=0, **kwargs):
        super().__init__(
            orientation="vertical", padding=dp(12), spacing=dp(10), **kwargs
        )

        self.app = MDApp.get_running_app()
        self.df = self.app.dataframe
        self.filtered_df = self.df.copy()
        self.ready = False

        saved_filters = self.app.cache_data.get("__current_filters__", {})
        self.filters_visible = saved_filters.get("filters_visible", True)

        self.filter_container = BoxLayout(
            orientation="vertical", size_hint_y=None
        )
        self.filter_container.bind(
            minimum_height=self.filter_container.setter("height")
        )

        self.filter_panel = BoxLayout(
            orientation="vertical", size_hint_y=None, spacing=dp(10)
        )
        self.filter_panel.bind(
            minimum_height=self.filter_panel.setter("height")
        )

        self.items_count_label = MDLabel(
            text=f"Знайдено товарів: {rows}",
            font_style="Subtitle2",
            theme_text_color="Secondary",
            size_hint_y=None,
            height=dp(25),
        )
        self.filter_panel.add_widget(self.items_count_label)

        spinner_style = {"bold": True, "font_size": "12sp"}

        # Фільтри Рядок 1
        self.filter_row1 = BoxLayout(
            orientation="horizontal",
            size_hint_y=None,
            height=dp(32),
            spacing=dp(6),
        )
        self.con_spinner = Spinner(text="Всі", values=[], **spinner_style)
        self.gz_spinner = Spinner(text="Всі", values=[], **spinner_style)

        self.filter_row1.add_widget(
            MDLabel(
                text="Конк:",
                size_hint_x=None,
                width=dp(32),
                font_style="Caption",
            )
        )
        self.filter_row1.add_widget(self.con_spinner)
        self.filter_row1.add_widget(
            MDLabel(
                text="ГЗ:", size_hint_x=None, width=dp(25), font_style="Caption"
            )
        )
        self.filter_row1.add_widget(self.gz_spinner)
        self.filter_panel.add_widget(self.filter_row1)

        # Фільтри Рядок 2
        self.filter_row2 = BoxLayout(
            orientation="horizontal",
            size_hint_y=None,
            height=dp(32),
            spacing=dp(6),
        )
        self.seg_spinner = Spinner(text="Всі", values=[], **spinner_style)
        self.tm_spinner = Spinner(text="Всі", values=[], **spinner_style)

        self.filter_row2.add_widget(
            MDLabel(
                text="Сегм:",
                size_hint_x=None,
                width=dp(32),
                font_style="Caption",
            )
        )
        self.filter_row2.add_widget(self.seg_spinner)
        self.filter_row2.add_widget(
            MDLabel(
                text="ТМ:", size_hint_x=None, width=dp(25), font_style="Caption"
            )
        )
        self.filter_row2.add_widget(self.tm_spinner)
        self.filter_panel.add_widget(self.filter_row2)

        self.con_spinner.text = saved_filters.get("con", "Всі")
        self.gz_spinner.text = saved_filters.get("gz", "Всі")
        self.seg_spinner.text = saved_filters.get("seg", "Всі")
        self.tm_spinner.text = saved_filters.get("tm", "Всі")

        self.con_spinner.bind(text=self.on_filter_change)
        self.gz_spinner.bind(text=self.on_filter_change)
        self.seg_spinner.bind(text=self.on_filter_change)
        self.tm_spinner.bind(text=self.on_filter_change)

        # Сортування
        self.sort_container = BoxLayout(
            orientation="horizontal",
            size_hint_y=None,
            height=dp(32),
            spacing=dp(6),
        )
        self.sort_spinner = Spinner(
            text=saved_filters.get("sort", "Без сортування"),
            values=[
                "Без сортування",
                "Назва: від А до Я",
                "Назва: від Я до А",
                "Наша ціна: від меншої до більшої",
                "Наша ціна: від більшої до меншої",
                "Різниця ціни %: спочатку наибольша",
                "Різниця ціни %: спочатку найменша",
            ],
            bold=True,
            font_size="12sp",
        )
        self.sort_spinner.bind(text=self.on_filter_change)
        self.sort_container.add_widget(
            MDLabel(
                text="Сорт:",
                size_hint_x=None,
                width=dp(40),
                font_style="Caption",
            )
        )
        self.sort_container.add_widget(self.sort_spinner)
        self.filter_panel.add_widget(self.sort_container)

        # Пошук
        search_wrap = BoxLayout(
            size_hint_y=None, height=dp(65), padding=[0, dp(5), 0, dp(5)]
        )
        search_anchor = AnchorLayout(anchor_x="center", anchor_y="center")

        self.search_field = MDTextField(
            hint_text="Пошук за назвою товару...",
            text=saved_filters.get("search", ""),
            mode="line",
            size_hint=(1, None),
            height=dp(56),
        )
        self.search_field.bind(text=self.on_filter_change)

        search_anchor.add_widget(self.search_field)
        search_wrap.add_widget(search_anchor)
        self.filter_panel.add_widget(search_wrap)

        if self.filters_visible:
            self.filter_container.add_widget(self.filter_panel)
        else:
            self.filter_container.height = 0
        self.add_widget(self.filter_container)

        # --- 2. КНОПКА-СТРІЛОЧКА ---
        self.toggle_anchor = AnchorLayout(
            size_hint_y=None,
            height=dp(24),
            anchor_x="center",
            anchor_y="center",
        )
        self.toggle_btn = MDIconButton(
            icon="chevron-up" if self.filters_visible else "chevron-down",
            theme_text_color="Custom",
            text_color=[0.5, 0.5, 0.5, 1],
            on_release=self.toggle_filters,
        )
        self.toggle_anchor.add_widget(self.toggle_btn)
        self.add_widget(self.toggle_anchor)

        # --- 3. СКРОЛЛ ТА ЕКСПОРТ ---
        self.scroll = ScrollView(size_hint=(1, 1))
        self.container = BoxLayout(
            orientation="vertical",
            spacing=dp(14),
            padding=[0, dp(8), 0, dp(8)],
            size_hint_y=None,
        )
        self.container.bind(minimum_height=self.container.setter("height"))
        self.scroll.add_widget(self.container)
        self.add_widget(self.scroll)

        self.export_btn = MDRaisedButton(
            text="ЕКСПОРТУВАТИ ЗМІНИ В XLSX",
            size_hint=(1, None),
            height=dp(60),
            md_bg_color=[0.1, 0.6, 0.3, 1],
            on_release=lambda x: self.app.save_to_original_excel(),
        )
        self.add_widget(self.export_btn)

        self.update_spinners(self.df)
        self.ready = True
        self.refresh_from_df()

    def toggle_filters(self, instance):
        if self.filters_visible:
            self.filter_container.clear_widgets()
            self.filter_container.height = 0
            self.toggle_btn.icon = "chevron-down"
            self.filters_visible = False
        else:
            self.filter_container.add_widget(self.filter_panel)
            self.toggle_btn.icon = "chevron-up"
            self.filters_visible = True

        self.save_all_filters_state()

    def on_filter_change(self, instance, text):
        if not self.ready:
            return
        self.save_all_filters_state()
        self.refresh_from_df()

    def save_all_filters_state(self):
        filter_state = {
            "con": self.con_spinner.text,
            "gz": self.gz_spinner.text,
            "seg": self.seg_spinner.text,
            "tm": self.tm_spinner.text,
            "sort": self.sort_spinner.text,
            "search": self.search_field.text,
            "filters_visible": self.filters_visible,
        }
        self.app.save_change_to_service_file(
            "__current_filters__", "data", filter_state
        )

    def clean_list(self, series):
        return sorted(
            [
                str(x)
                for x in series.unique().tolist()
                if str(x).strip() not in ["", "nan", "None"]
            ]
        )

    def update_spinners(self, df):
        self.con_spinner.values = ["Всі"] + self.clean_list(df["Конкурент"])
        self.gz_spinner.values = ["Всі"] + self.clean_list(df["ГЗ"])
        self.seg_spinner.values = ["Всі"] + self.clean_list(df["Сегмент"])
        self.tm_spinner.values = ["Всі"] + self.clean_list(df["ТМ"])

    def refresh_from_df(self):
        self.ready = False
        df = self.get_filtered_df()

        con, gz, seg, tm = (
            self.con_spinner.text,
            self.gz_spinner.text,
            self.seg_spinner.text,
            self.tm_spinner.text,
        )
        # Спочатку фільтруємо основний датафрейм (як у вас було)
        self.con_spinner.values = ["Всі"] + self.clean_list(
            self.get_filtered_df("con")["Конкурент"]
        )
        self.gz_spinner.values = ["Всі"] + self.clean_list(
            self.get_filtered_df("gz")["ГЗ"]
        )
        self.seg_spinner.values = ["Всі"] + self.clean_list(
            self.get_filtered_df("seg")["Сегмент"]
        )
        self.tm_spinner.values = ["Всі"] + self.clean_list(
            self.get_filtered_df("tm")["ТМ"]
        )

        search_query = self.search_field.text.strip().lower()
        if search_query:
            df = df[
                df["Товар"].str.lower().str.contains(search_query, na=False)
            ]

        # ... [Тут залишається ваша логіка сортування, яку я не чіпаю] ...
        sort_mode = self.sort_spinner.text
        if sort_mode == "Назва: від А до Я":
            df = df.sort_values(
                by="Товар", ascending=True, key=lambda col: col.str.lower()
            )
        elif sort_mode == "Назва: від Я до А":
            df = df.sort_values(
                by="Товар", ascending=False, key=lambda col: col.str.lower()
            )
        elif sort_mode == "Наша ціна: від меншої до більшої":
            df["_temp_our_price"] = (
                df["Розн. цена"].apply(extract_number).fillna(0.0)
            )
            df = df.sort_values(by="_temp_our_price", ascending=True).drop(
                columns=["_temp_our_price"]
            )
        elif "Наша ціна: від більшої до меншої" in sort_mode:
            df["_temp_our_price"] = (
                df["Розн. цена"].apply(extract_number).fillna(0.0)
            )
            df = df.sort_values(by="_temp_our_price", ascending=False).drop(
                columns=["_temp_our_price"]
            )
        elif "Різниця ціни %" in sort_mode:

            def calc_diff_pct(row):
                p1 = extract_number(row.get("Розн. цена"))
                p2 = extract_number(row.get("Цена конкурента"))
                if p1 and p2 and p1 > 0:
                    return abs(p2 - p1) / p1 * 100
                return -1.0

            df["_temp_diff_pct"] = df.apply(calc_diff_pct, axis=1)
            if sort_mode == "Різниця ціни %: спочатку найбільша":
                df = df.sort_values(by="_temp_diff_pct", ascending=False).drop(
                    columns=["_temp_diff_pct"]
                )
            elif sort_mode == "Різниця ціни %: спочатку найменша":
                df_valid = df[df["_temp_diff_pct"] >= 0].sort_values(
                    by="_temp_diff_pct", ascending=True
                )
                df_invalid = df[df["_temp_diff_pct"] < 0]
                df = pd.concat([df_valid, df_invalid]).drop(
                    columns=["_temp_diff_pct"]
                )

        self.filtered_df = df
        self.items_count_label.text = f"Знайдено товарів: {len(df)}"

        # ОСЬ ТУТ ДОДАНО ОНОВЛЕННЯ ДЛЯ ТМ:
        # Якщо фільтр не обраний (Всі), оновлюємо його список на основі поточної вибірки
        if con == "Всі":
            self.con_spinner.values = ["Всі"] + self.clean_list(df["Конкурент"])
        if gz == "Всі":
            self.gz_spinner.values = ["Всі"] + self.clean_list(df["ГЗ"])
        if seg == "Всі":
            self.seg_spinner.values = ["Всі"] + self.clean_list(df["Сегмент"])
        if tm == "Всі":
            self.tm_spinner.values = ["Всі"] + self.clean_list(df["ТМ"])

        self.ready = True
        self.build_list()

    def build_list(self):
        self.container.clear_widgets()
        for i, row in self.filtered_df.head(150).iterrows():
            card = ProductCard(
                i,
                str(row.get("ТМ", "")),
                str(row.get("Товар", "")),
                str(row.get("Aртикул", row.get("Артикул", ""))),
                str(row.get("Розн. цена", "")),
                str(row.get("Цена конкурента", "")),
                str(row.get("Конкурент", "")),
                str(
                    row.get(
                        "Alternat. цена",
                        row.get(
                            "Aktcyonnaya tsena", row.get("Акционная цена", "")
                        ),
                    )
                ),
                str(row.get("Акция", "")),
                str(row.get("Примечание", "")),
                str(row.get("Ед.изм.", "")),
                str(row.get("Сегмент", "")),
            )
            self.container.add_widget(card)

    def get_filtered_df(self, exclude=None):
        df = self.app.dataframe.copy().fillna("")

        con, gz, seg, tm = (
            self.con_spinner.text,
            self.gz_spinner.text,
            self.seg_spinner.text,
            self.tm_spinner.text,
        )

        if exclude != "con" and con != "Всі":
            df = df[df["Конкурент"] == con]

        if exclude != "gz" and gz != "Всі":
            df = df[df["ГЗ"] == gz]

        if exclude != "seg" and seg != "Всі":
            df = df[df["Сегмент"] == seg]

        if exclude != "tm" and tm != "Всі":
            df = df[df["ТМ"] == tm]

        search_query = self.search_field.text.strip().lower()
        if search_query:
            df = df[
                df["Товар"].str.lower().str.contains(search_query, na=False)
            ]

        return df


# ---------- ГОЛОВНИЙ ROOT МЕНЕДЖЕР ----------
class Root(MDNavigationLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.screen_manager = MDScreenManager()
        self.main_screen = MDScreen(name="main_screen")
        self.screen_layout = BoxLayout(orientation="vertical")

        self.toolbar = MDTopAppBar(
            title="Інструмент моніторингу",
            elevation=4,
            left_action_items=[
                ["menu", lambda x: self.nav_drawer.set_state("open")]
            ],
        )
        self.screen_layout.add_widget(self.toolbar)

        self.content_area = BoxLayout()
        self.screen_layout.add_widget(self.content_area)

        self.main_screen.add_widget(self.screen_layout)
        self.screen_manager.add_widget(self.main_screen)
        self.add_widget(self.screen_manager)

        self.nav_drawer = MDNavigationDrawer(radius=(0, 16, 16, 0))
        self.drawer_content = BoxLayout(
            orientation="vertical", padding=dp(12), spacing=dp(10)
        )

        self.drawer_content.add_widget(
            MDLabel(
                text="Останні файли:",
                font_style="Subtitle1",
                bold=True,
                size_hint_y=None,
                height=dp(35),
            )
        )

        self.recent_files_list = MDList()
        scroll_files = ScrollView()
        scroll_files.add_widget(self.recent_files_list)
        self.drawer_content.add_widget(scroll_files)

        footer_box = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            height=dp(160),
            spacing=dp(0),
        )
        footer_box.add_widget(MDSeparator())

        self.btn_open_file = OneLineIconListItem(
            text="Вибрати інший файл",
            on_release=lambda x: self.go_to_file_screen(),
        )
        self.btn_open_file.add_widget(IconLeftWidget(icon="folder"))
        footer_box.add_widget(self.btn_open_file)

        self.theme_btn = OneLineIconListItem(
            text="Змінити тему", on_release=self.toggle_theme
        )
        self.theme_btn.add_widget(IconLeftWidget(icon="theme-light-dark"))
        footer_box.add_widget(self.theme_btn)

        self.btn_about = OneLineIconListItem(
            text="Про додаток", on_release=self.show_about_dialog
        )
        self.btn_about.add_widget(IconLeftWidget(icon="information"))
        footer_box.add_widget(self.btn_about)

        footer_box.add_widget(
            MDLabel(
                text="© 2026, Vitaliy Shkarupa",
                font_style="Caption",
                theme_text_color="Secondary",
                halign="center",
                size_hint_y=None,
                height=dp(30),
            )
        )

        self.drawer_content.add_widget(footer_box)
        self.nav_drawer.add_widget(self.drawer_content)
        self.add_widget(self.nav_drawer)

        self.show_file_screen()

    def show_file_screen(self):
        self.content_area.clear_widgets()
        self.content_area.add_widget(FileScreen())
        self.toolbar.title = "Оберіть xlsx файл"
        self.toolbar.left_action_items = [
            ["menu", lambda x: self.nav_drawer.set_state("open")]
        ]
        self.toolbar.right_action_items = [["", lambda x: None]]

    def show_main_screen(self, rows):
        self.content_area.clear_widgets()
        self.content_area.add_widget(MainScreen(rows))

        app = MDApp.get_running_app()

        filename = os.path.basename(app.source_file_path)
        filename_no_ext = os.path.splitext(filename)[0]

        self.toolbar.title = f"Моніторинг — {filename_no_ext}"

        self.toolbar.left_action_items = [
            ["menu", lambda x: self.nav_drawer.set_state("open")]
        ]

        self.toolbar.right_action_items = [["", lambda x: None]]

    def refresh_recent_files_menu(self):
        self.recent_files_list.clear_widgets()
        app = MDApp.get_running_app()
        recent_paths = app.global_settings.get("recent_files", [])
        if not recent_paths:
            self.recent_files_list.add_widget(
                OneLineListItem(
                    text="Немає нещодавніх файлів", theme_text_color="Hint"
                )
            )
            return
        for path in recent_paths:
            file_name = os.path.basename(path)
            item = OneLineIconListItem(
                text=file_name,
                on_release=lambda x, p=path: self.load_recent_file(p),
            )
            item.add_widget(IconLeftWidget(icon="file-excel"))
            self.recent_files_list.add_widget(item)

    def load_recent_file(self, path):
        self.nav_drawer.set_state("close")
        if os.path.exists(path):
            app = MDApp.get_running_app()
            app.source_file_path = path
            app.load_data_from_source()
        else:
            toast("Помилка: Файл більше не існує!")

    def show_about_dialog(self, instance):
        dialog = MDDialog(
            title="Про додаток",
            text="Додаток призначений для швидкого моніторингу цін конкурентів.\n\n"
            "• Виберіть Excel файл для завантаження списку.\n"
            "• Вводьте ціни конкурентів прямо в полі картки товару.\n"
            "• Колір підсвічування змінюється залежно від різниці у відсотках.\n"
            "• Кнопка 'Експортувати' перезаписує зміни безпосередньо у вихідний файл Excel.\n"
            "• Додаток автоматично кешує ваші проміжні зміни.",
            buttons=[
                MDFlatButton(text="OK", on_release=lambda x: dialog.dismiss())
            ],
        )
        dialog.open()

    def refresh_recent_files_menu(self):
        self.recent_files_list.clear_widgets()

        app = MDApp.get_running_app()
        recent_paths = app.global_settings.get("recent_files", [])

        if not recent_paths:
            self.recent_files_list.add_widget(
                OneLineListItem(text="Список порожній")
            )
            return

        for path in recent_paths:
            file_name = os.path.basename(path)

            row = BoxLayout(size_hint_y=None, height=dp(48))

            item = OneLineIconListItem(
                text=file_name,
                on_release=lambda x, p=path: self.load_recent_file(p),
            )
            item.add_widget(IconLeftWidget(icon="file-excel"))

            delete_btn = MDIconButton(
                icon="trash-can",
                theme_text_color="Custom",
                text_color=[1, 0.3, 0.3, 1],
                on_release=lambda x, p=path: self.remove_recent_file(p),
            )

            row.add_widget(item)
            row.add_widget(delete_btn)
            self.recent_files_list.add_widget(row)

    def remove_recent_file(self, path):
        app = MDApp.get_running_app()
        recent = app.global_settings.get("recent_files", [])

        if path in recent:
            recent.remove(path)
            app.global_settings["recent_files"] = recent

            try:
                with open(app.global_settings_path, "w", encoding="utf-8") as f:
                    json.dump(
                        app.global_settings, f, ensure_ascii=False, indent=2
                    )
            except Exception:
                pass

            self.refresh_recent_files_menu()

    def go_to_file_screen(self):
        self.nav_drawer.set_state("close")
        self.show_file_screen()

    def toggle_theme(self, instance):
        app = MDApp.get_running_app()
        app.toggle_theme(instance)


# ---------- ДОДАТОК ----------
class PriceApp(MDApp):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dataframe = pd.DataFrame()
        self.source_file_path = ""
        self.last_mtime = 0.0
        self.disable_sync = False
        self.cache_data = {}
        self.global_settings = {}

    def build(self):
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Dark"
        self.load_global_settings()

        self.theme_cls.theme_style = self.global_settings.get("theme", "Dark")

        root_ui = Root()
        Clock.schedule_once(lambda dt: root_ui.refresh_recent_files_menu(), 0.2)
        Clock.schedule_interval(self.check_file_sync, 5.0)
        return root_ui

    @property
    def global_settings_path(self):
        return os.path.join(self.user_data_dir, "global_settings.json")

    def load_global_settings(self):
        if os.path.exists(self.global_settings_path):
            try:
                with open(
                    self.global_settings_path, "r", encoding="utf-8"
                ) as f:
                    self.global_settings = json.load(f)
            except Exception:
                self.global_settings = {"recent_files": []}
        else:
            self.global_settings = {"recent_files": []}

    def save_file_to_recent(self, path):
        if not path:
            return
        recent = self.global_settings.get("recent_files", [])
        if path in recent:
            recent.remove(path)
        recent.insert(0, path)
        self.global_settings["recent_files"] = recent[:5]
        try:
            with open(self.global_settings_path, "w", encoding="utf-8") as f:
                json.dump(self.global_settings, f, ensure_ascii=False, indent=2)
            self.root.refresh_recent_files_menu()
        except Exception as e:
            print("Помилка збереження історії файлів:", e)

    @property
    def service_file_path(self):
        if not self.source_file_path:
            return os.path.join(self.user_data_dir, "default_cache.json")
        file_hash = hashlib.md5(
            self.source_file_path.encode("utf-8")
        ).hexdigest()[:10]
        base_name = os.path.basename(self.source_file_path)
        name_part, _ = os.path.splitext(base_name)
        return os.path.join(
            self.user_data_dir, f"cache_{name_part}_{file_hash}.json"
        )

    def load_cache_for_current_file(self):
        cache_path = self.service_file_path
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    self.cache_data = json.load(f)
            except Exception:
                self.cache_data = {}
        else:
            self.cache_data = {}

    def save_change_to_service_file(self, row_idx, column_name, value):
        idx_str = str(row_idx)
        if idx_str == "__current_filters__":
            self.cache_data[idx_str] = value
        else:
            if idx_str not in self.cache_data:
                self.cache_data[idx_str] = {}
            self.cache_data[idx_str][column_name] = value
        try:
            with open(self.service_file_path, "w", encoding="utf-8") as f:
                json.dump(self.cache_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            pass

    def apply_cache_to_dataframe(self):
        for idx_str, fields in self.cache_data.items():
            if idx_str == "__current_filters__":
                continue
            try:
                idx = int(idx_str)
                if idx < len(self.dataframe):
                    for col, val in fields.items():
                        clean_col = "Акция" if col in ["Акция"] else col
                        self.dataframe.at[idx, clean_col] = val
            except ValueError:
                continue

    def load_data_from_source(self):
        if not os.path.exists(self.source_file_path):
            return
        try:
            self.last_mtime = os.path.getmtime(self.source_file_path)
            wb = load_workbook(self.source_file_path, data_only=True)
            sheet = wb.active
            headers, data, first = [], [], True
            for row in sheet.iter_rows(values_only=True):
                if first:
                    headers = [
                        str(h).strip() if h is not None else "" for h in row
                    ]
                    first = False
                    continue
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                data.append(row_dict)
            self.dataframe = pd.DataFrame(data)

            required_cols = [
                "ТМ",
                "Товар",
                "Артикул",
                "Розн. цена",
                "Ед.изм.",
                "ЦенаBox",
                "Цена competitor",
                "Цена конкурента",
                "Акционная цена",
                "Акция",
                "Примечание",
                "Конкурент",
                "ГЗ",
                "Сегмент",
            ]
            for col in required_cols:
                if col not in self.dataframe.columns:
                    self.dataframe[col] = ""

            self.load_cache_for_current_file()
            self.apply_cache_to_dataframe()
            self.save_file_to_recent(self.source_file_path)

            if isinstance(self.root.content_area.children[0], MainScreen):
                self.root.content_area.children[0].refresh_from_df()
            else:
                self.root.show_main_screen(len(self.dataframe))
        except Exception as e:
            pass

    def save_to_original_excel(self):
        if self.dataframe.empty or not self.source_file_path:
            toast("Помилка: Немає даних для експорту!")
            return
        self.disable_sync = True
        try:
            toast("Експортуємо зміни в xlsx... Зачекайте.")
            saved_filters = self.cache_data.get("__current_filters__", {})

            if "Акция" in self.dataframe.columns:
                self.dataframe = self.dataframe.drop(columns=["Акция"])

            self.dataframe.to_excel(self.source_file_path, index=False)
            self.last_mtime = os.path.getmtime(self.source_file_path)

            self.cache_data = {}
            if saved_filters:
                self.cache_data["__current_filters__"] = saved_filters
                try:
                    with open(
                        self.service_file_path, "w", encoding="utf-8"
                    ) as f:
                        json.dump(
                            self.cache_data, f, ensure_ascii=False, indent=2
                        )
                except Exception:
                    pass
            elif os.path.exists(self.service_file_path):
                os.remove(self.service_file_path)
            toast("Експорт файла успішний!")
        except Exception as e:
            toast("Помилка: Закрийте xlsx файл в інших додатках!")
        self.disable_sync = False

    def check_file_sync(self, dt):
        if (
            self.disable_sync
            or not self.source_file_path
            or not os.path.exists(self.source_file_path)
        ):
            return
        try:
            current_mtime = os.path.getmtime(self.source_file_path)
            if current_mtime > self.last_mtime:
                self.last_mtime = current_mtime
                self.load_data_from_source()
                toast("Зовнішній xlsx файл оновився. Синхронізуємо дані...")
        except Exception as e:
            pass

    def toggle_theme(self, instance):
        if self.theme_cls.theme_style == "Dark":
            self.theme_cls.theme_style = "Light"
            self.global_settings["theme"] = "Light"
        else:
            self.theme_cls.theme_style = "Dark"
            self.global_settings["theme"] = "Dark"

        try:
            with open(self.global_settings_path, "w", encoding="utf-8") as f:
                json.dump(self.global_settings, f, ensure_ascii=False, indent=2)
        except Exception:
            pass


if __name__ == "__main__":
    PriceApp().run()